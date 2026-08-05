import json
import os
import tempfile
from pathlib import Path
from typing import Any

import requests
from fastapi import UploadFile

WORKSPACE = Path(__file__).resolve().parents[1]
API_URL = "http://127.0.0.1:8001"

FILE_DIR = WORKSPACE / "tests" / "validation_files"
FILE_DIR.mkdir(parents=True, exist_ok=True)

TOPOLOGY = {
    "assets": [
        {"id": "Internet", "name": "Internet", "type": "device", "zone": "Internet"},
        {"id": "Corporate_Network", "name": "Corporate Network", "type": "device", "zone": "Corporate"},
        {"id": "DMZ", "name": "DMZ", "type": "device", "zone": "DMZ"},
        {"id": "Firewall", "name": "Firewall", "type": "device", "zone": "DMZ"},
        {"id": "VPN_Gateway", "name": "VPN Gateway", "type": "device", "zone": "Corporate"},
        {"id": "Remote_Access_Server", "name": "Remote Access Server", "type": "device", "zone": "Corporate"},
        {"id": "Engineering_Workstation", "name": "Engineering Workstation", "type": "device", "zone": "Corporate"},
        {"id": "SCADA_Server", "name": "SCADA Server", "type": "device", "zone": "Level 3"},
        {"id": "Historian", "name": "Historian", "type": "device", "zone": "Level 3"},
        {"id": "HMI", "name": "HMI", "type": "device", "zone": "Level 2"},
        {"id": "Domain_Controller", "name": "Domain Controller", "type": "device", "zone": "Corporate"},
        {"id": "Patch_Management_Server", "name": "Patch Management Server", "type": "device", "zone": "Corporate"},
        {"id": "OPC_Server", "name": "OPC Server", "type": "device", "zone": "Level 2"},
        {"id": "PLC_1", "name": "PLC 1", "type": "device", "zone": "Level 1"},
        {"id": "PLC_2", "name": "PLC 2", "type": "device", "zone": "Level 1"},
        {"id": "RTU", "name": "RTU", "type": "device", "zone": "Level 1"},
        {"id": "Remote_Site", "name": "Remote Site", "type": "device", "zone": "Remote"},
        {"id": "Level_3_Switch", "name": "Level 3 Switch", "type": "device", "zone": "Level 3"},
        {"id": "Level_2_Switch", "name": "Level 2 Switch", "type": "device", "zone": "Level 2"},
        {"id": "Sensor", "name": "Sensor", "type": "physical", "zone": "Level 0"},
        {"id": "Actuator", "name": "Actuator", "type": "physical", "zone": "Level 0"},
        {"id": "Safety_PLC", "name": "Safety PLC", "type": "device", "zone": "Level 1"},
    ],
    "relationships": [
        {"source": "Internet", "target": "Firewall", "type": "connects-to", "protocol": "http", "trust_level": "low"},
        {"source": "Firewall", "target": "DMZ", "type": "connects-to", "protocol": "http", "trust_level": "medium"},
        {"source": "DMZ", "target": "VPN_Gateway", "type": "connects-to", "protocol": "ipsec", "trust_level": "medium"},
        {"source": "VPN_Gateway", "target": "Corporate_Network", "type": "connects-to", "protocol": "ipsec", "trust_level": "high"},
        {"source": "Corporate_Network", "target": "Engineering_Workstation", "type": "connects-to", "protocol": "rdp", "trust_level": "high"},
        {"source": "Corporate_Network", "target": "Domain_Controller", "type": "connects-to", "protocol": "ldap", "trust_level": "high"},
        {"source": "Corporate_Network", "target": "Patch_Management_Server", "type": "connects-to", "protocol": "wsman", "trust_level": "high"},
        {"source": "Corporate_Network", "target": "SCADA_Server", "type": "connects-to", "protocol": "opc-ua", "trust_level": "medium"},
        {"source": "SCADA_Server", "target": "Historian", "type": "connects-to", "protocol": "opc-ua", "trust_level": "medium"},
        {"source": "SCADA_Server", "target": "HMI", "type": "connects-to", "protocol": "opc-ua", "trust_level": "medium"},
        {"source": "SCADA_Server", "target": "OPC_Server", "type": "connects-to", "protocol": "opc-ua", "trust_level": "medium"},
        {"source": "OPC_Server", "target": "PLC_1", "type": "connects-to", "protocol": "modbus", "trust_level": "low"},
        {"source": "OPC_Server", "target": "PLC_2", "type": "connects-to", "protocol": "profinet", "trust_level": "low"},
        {"source": "PLC_1", "target": "Sensor", "type": "connects-to", "protocol": "modbus", "trust_level": "low"},
        {"source": "PLC_2", "target": "Actuator", "type": "connects-to", "protocol": "ethernet/ip", "trust_level": "low"},
        {"source": "PLC_1", "target": "RTU", "type": "connects-to", "protocol": "dnp3", "trust_level": "low"},
        {"source": "RTU", "target": "Remote_Site", "type": "connects-to", "protocol": "dnp3", "trust_level": "low"},
        {"source": "SCADA_Server", "target": "Safety_PLC", "type": "connects-to", "protocol": "profinet", "trust_level": "medium"},
        {"source": "Safety_PLC", "target": "Sensor", "type": "connects-to", "protocol": "modbus", "trust_level": "low"},
    ],
}


OUTPUT_LOG = WORKSPACE / "tests" / "validation_report.json"


def write_json(path: Path, data: Any) -> None:
    with path.open("w", encoding="utf-8") as handle:
        json.dump(data, handle, indent=2)


def write_csv(path: Path, rows: list[list[str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        for row in rows:
            handle.write(",".join(row) + "\n")


def write_excel(path: Path, rows: list[list[str]], sheet_name: str = "Sheet1") -> None:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(path)


def write_graphml(path: Path, topology: dict[str, Any]) -> None:
    import networkx as nx
    graph = nx.DiGraph()
    for asset in topology["assets"]:
        graph.add_node(asset["id"], **{k: v for k, v in asset.items() if k not in {"id"}})
    for rel in topology["relationships"]:
        graph.add_edge(rel["source"], rel["target"], **{k: v for k, v in rel.items() if k not in {"source", "target"}})
    nx.write_graphml(graph, str(path))


def write_xml(path: Path, topology: dict[str, Any]) -> None:
    import xml.etree.ElementTree as ET
    root = ET.Element("Topology")
    assets_elem = ET.SubElement(root, "Assets")
    for asset in topology["assets"]:
        asset_elem = ET.SubElement(assets_elem, "Asset")
        for k, v in asset.items():
            child = ET.SubElement(asset_elem, k)
            child.text = str(v)
    rels_elem = ET.SubElement(root, "Relationships")
    for rel in topology["relationships"]:
        rel_elem = ET.SubElement(rels_elem, "Relationship")
        for k, v in rel.items():
            child = ET.SubElement(rel_elem, k)
            child.text = str(v)
    tree = ET.ElementTree(root)
    tree.write(path, encoding="utf-8", xml_declaration=True)


def write_aml(path: Path, topology: dict[str, Any]) -> None:
    import xml.etree.ElementTree as ET
    root = ET.Element("AdditionalMarkupLanguage")
    system_unit = ET.SubElement(root, "SystemUnitClassLib")
    for asset in topology["assets"]:
        element = ET.SubElement(system_unit, "InternalElement", Name=asset["id"])
        for k, v in asset.items():
            if k in {"id", "name"}:
                continue
            attr = ET.SubElement(element, "Attribute", Name=k)
            attr.text = str(v)
    for rel in topology["relationships"]:
        connection = ET.SubElement(root, "Connection")
        for k, v in rel.items():
            child = ET.SubElement(connection, k.capitalize())
            child.text = str(v)
    tree = ET.ElementTree(root)
    tree.write(path, encoding="utf-8", xml_declaration=True)


def write_vsdx(path: Path) -> None:
    import zipfile

    content_types = """<?xml version='1.0' encoding='UTF-8'?>
<Types xmlns='http://schemas.openxmlformats.org/package/2006/content-types'>
  <Default Extension='rels' ContentType='application/vnd.openxmlformats-package.relationships+xml'/>
  <Default Extension='xml' ContentType='application/xml'/>
  <Override PartName='/docProps/app.xml' ContentType='application/vnd.openxmlformats-officedocument.extended-properties+xml'/>
  <Override PartName='/visio/document.xml' ContentType='application/vnd.ms-visio.document.main+xml'/>
  <Override PartName='/visio/_rels/document.xml.rels' ContentType='application/vnd.openxmlformats-package.relationships+xml'/>
  <Override PartName='/visio/pages/pages.xml' ContentType='application/vnd.ms-visio.pages+xml'/>
  <Override PartName='/visio/pages/_rels/pages.xml.rels' ContentType='application/vnd.openxmlformats-package.relationships+xml'/>
  <Override PartName='/visio/pages/page1.xml' ContentType='application/vnd.ms-visio.page+xml'/>
</Types>"""

    app_xml = """<?xml version='1.0' encoding='UTF-8'?>
<Properties xmlns='http://schemas.openxmlformats.org/officeDocument/2006/extended-properties' xmlns:vt='http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes'>
  <HeadingPairs>
    <vt:vector size='2'>
      <vt:variant><vt:lpstr>Pages</vt:lpstr></vt:variant>
      <vt:variant><vt:i4>1</vt:i4></vt:variant>
    </vt:vector>
  </HeadingPairs>
  <TitlesOfParts>
    <vt:vector size='1'>
      <vt:lpstr>Page-1</vt:lpstr>
    </vt:vector>
  </TitlesOfParts>
</Properties>"""

    document_xml = """<?xml version='1.0' encoding='UTF-8'?>
<Document xmlns='http://schemas.microsoft.com/office/visio/2012/main'/>
"""

    document_rels = """<?xml version='1.0' encoding='UTF-8'?>
<Relationships xmlns='http://schemas.openxmlformats.org/package/2006/relationships'/>
"""

    pages_xml = """<?xml version='1.0' encoding='UTF-8'?>
<Pages xmlns='http://schemas.microsoft.com/office/visio/2012/main'>
  <Page ID='1' Name='Page-1' NameU='Page-1'>
    <Rel r:id='rId1' xmlns:r='http://schemas.openxmlformats.org/officeDocument/2006/relationships'/>
  </Page>
</Pages>"""

    pages_rels = """<?xml version='1.0' encoding='UTF-8'?>
<Relationships xmlns='http://schemas.openxmlformats.org/package/2006/relationships'>
  <Relationship Id='rId1' Type='http://schemas.microsoft.com/visio/2010/relationships/page' Target='page1.xml'/>
</Relationships>"""

    page_xml = """<?xml version='1.0' encoding='UTF-8'?>
<PageContents xmlns='http://schemas.microsoft.com/office/visio/2012/main'>
  <Shapes>
"""
    for asset in TOPOLOGY['assets']:
        page_xml += f"    <Shape ID='{asset['id']}'><Text>asset,{asset['id']},{asset.get('type','')},{asset.get('zone','')}</Text></Shape>\n"
    for rel in TOPOLOGY['relationships']:
        page_xml += (
            f"    <Shape ID='{rel['source']}_{rel['target']}'><Text>relationship,{rel['source']},{rel['target']},{rel['type']},false,{rel.get('protocol','')},{rel.get('trust_level','')}</Text></Shape>\n"
        )
    page_xml += "  </Shapes>\n</PageContents>"

    with zipfile.ZipFile(path, 'w', compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr('[Content_Types].xml', content_types)
        archive.writestr('docProps/app.xml', app_xml)
        archive.writestr('visio/document.xml', document_xml)
        archive.writestr('visio/_rels/document.xml.rels', document_rels)
        archive.writestr('visio/pages/pages.xml', pages_xml)
        archive.writestr('visio/pages/_rels/pages.xml.rels', pages_rels)
        archive.writestr('visio/pages/page1.xml', page_xml)


def write_vsd(path: Path) -> None:
    # Legacy VSD cannot be generated without Visio; use a placeholder extension with XML content for basic path coverage.
    path.write_text("<visio-legacy>Legacy VSD placeholder</visio-legacy>", encoding="utf-8")


def make_files() -> dict[str, Path]:
    paths: dict[str, Path] = {}
    write_json(FILE_DIR / "topology.json", TOPOLOGY)
    paths["json"] = FILE_DIR / "topology.json"

    csv_rows: list[list[str]] = [["id","name","type","zone"]]
    csv_rows += [[asset["id"], asset["name"], asset["type"], asset["zone"]] for asset in TOPOLOGY["assets"]]
    csv_rows += [[]]
    csv_rows += [["source","target","type","protocol","trust_level"]]
    csv_rows += [[rel["source"], rel["target"], rel["type"], rel["protocol"], rel["trust_level"]] for rel in TOPOLOGY["relationships"]]
    write_csv(FILE_DIR / "topology.csv", csv_rows)
    paths["csv"] = FILE_DIR / "topology.csv"

    from openpyxl import Workbook
    wb = Workbook()
    ws_assets = wb.active
    ws_assets.title = "Assets"
    ws_assets.append(["id","name","type","zone"])
    for asset in TOPOLOGY["assets"]:
        ws_assets.append([asset["id"], asset["name"], asset["type"], asset["zone"]])
    ws_rels = wb.create_sheet("Relationships")
    ws_rels.append(["source","target","type","protocol","trust_level"])
    for rel in TOPOLOGY["relationships"]:
        ws_rels.append([rel["source"], rel["target"], rel["type"], rel["protocol"], rel["trust_level"]])
    wb.save(FILE_DIR / "topology.xlsx")
    paths["xlsx"] = FILE_DIR / "topology.xlsx"

    write_graphml(FILE_DIR / "topology.graphml", TOPOLOGY)
    paths["graphml"] = FILE_DIR / "topology.graphml"
    write_xml(FILE_DIR / "topology.xml", TOPOLOGY)
    paths["xml"] = FILE_DIR / "topology.xml"
    write_aml(FILE_DIR / "topology.aml", TOPOLOGY)
    paths["aml"] = FILE_DIR / "topology.aml"
    write_vsdx(FILE_DIR / "topology.vsdx")
    paths["vsdx"] = FILE_DIR / "topology.vsdx"
    write_vsd(FILE_DIR / "topology.vsd")
    paths["vsd"] = FILE_DIR / "topology.vsd"
    return paths


def upload_file(path: Path) -> dict[str, Any]:
    with path.open("rb") as handle:
        files = {"file": (path.name, handle, "application/octet-stream")}
        response = requests.post(f"{API_URL}/upload-topology-file", files=files)
    return {
        "path": str(path),
        "status_code": response.status_code,
        "response": response.json() if response.headers.get("content-type",""
                                                              ).startswith("application/json") else response.text,
    }


def analyze_topology(topology: dict[str, Any]) -> dict[str, Any]:
    payload = {"topology": topology, "evidence": []}
    response = requests.post(f"{API_URL}/analyze", json=payload)
    return {
        "status_code": response.status_code,
        "response": response.json() if response.headers.get("content-type",""
                                                              ).startswith("application/json") else response.text,
    }


def main() -> None:
    paths = make_files()
    results: dict[str, Any] = {"uploads": {}, "analyses": {}, "summary": {}}
    for fmt, path in paths.items():
        upload = upload_file(path)
        results["uploads"][fmt] = upload
        if upload["status_code"] == 200 and isinstance(upload["response"], dict):
            topology = upload["response"].get("topology")
            if topology:
                analysis = analyze_topology(topology)
                results["analyses"][fmt] = analysis
    write_json(OUTPUT_LOG, results)
    print(json.dumps(results, indent=2))


if __name__ == "__main__":
    main()
